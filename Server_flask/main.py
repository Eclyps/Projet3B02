from flask import Flask, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import matplotlib.pyplot as plt
from fpdf import FPDF
import os
from datetime import datetime

app = Flask(__name__)

# Var de mémoire
liste_alerte = []
dic_vitesse = {}
dic_distance = {}
dic_vitesse_max = {}
dic_distance_sprint = {}
message = ""

# Nom du fichier Excel
EXCEL_FILE = "log_donnees.xlsx"

# Liste des colonnes dans excel
COLUMNS = ["Timestamp", "Vitesse", "Distance", "Vitessemax", "Distancesprint", "Alerte"]

#p Pour créer le fichier excel
def init_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "default"
        wb.save(EXCEL_FILE)

# Fonction pour écrire dans la feuille de l'ID
def write_event(post_id, data_type, value):
    wb = load_workbook(EXCEL_FILE)

    sheet_name = f"ID_{post_id}"

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(COLUMNS)
    else:
        ws = wb[sheet_name]

    # Trouver l'index de la colonne
    col_index = next((i for i, col in enumerate(COLUMNS) if col.lower() == data_type.lower()), None)
    if col_index is None:
        print(f"Type {data_type} non trouvé dans les colonnes!")
        wb.save(EXCEL_FILE)
        return
    col_index += 1  # Excel compte à partir de 1

    # Chercher une ligne disponible
    found_row = None
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=col_index).value
        if cell_value in (None, ""):
            found_row = row
            break

    # Si aucune ligne vide trouvée on en crée une
    if found_row is None:
        found_row = ws.max_row + 1
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=found_row, column=1, value=timestamp)

        # Mettre "-" dans Alerte par défaut
        alerte_col_index = COLUMNS.index("Alerte") + 1
        ws.cell(row=found_row, column=alerte_col_index, value="-")

    # Ajoute le timestamp s'il manque
    if ws.cell(row=found_row, column=1).value in (None, ""):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=found_row, column=1, value=timestamp)

    # Écrit la valeur dans la cellule cible
    ws.cell(row=found_row, column=col_index, value=value)

    # Si c'est une alerte, on met en rouge
    if data_type.lower() == "alerte" and str(value).lower() == "alerte":
        cell = ws.cell(row=found_row, column=col_index)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell.fill = red_fill

    wb.save(EXCEL_FILE)


init_excel_file()

OUTPUT_FOLDER = "pdf_reports"

# Lecture des données du fichier excel pour une joueuese
def read_data_from_excel(post_id):
    wb = load_workbook(EXCEL_FILE)
    sheet_name = f"ID_{post_id}"

    if sheet_name not in wb.sheetnames:
        print(f"Feuille {sheet_name} non trouvée !")
        return None

    df = pd.DataFrame(wb[sheet_name].values)
    df.columns = df.iloc[0]  # Premières lignes = headers
    df = df.drop(0).reset_index(drop=True)

    # On convertit Timestamp en datetime si nécessaire
    if 'Timestamp' in df.columns:
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])

    return df

#Fait la courbe vitesse
def plot_speed(df, post_id):
    plt.figure(figsize=(8, 4))
    plt.plot(df['Timestamp'], df['Vitesse'].astype(float), marker='o', color='blue')
    plt.title(f'Vitesse en fonction du temps (ID {post_id})')
    plt.xlabel('Temps')
    plt.ylabel('Vitesse (km/h)')
    plt.xticks(rotation=45)
    plt.tight_layout()
    graph_path = f"speed_time_{post_id}.png"
    plt.savefig(graph_path)
    plt.close()

    return graph_path

#Proportion sprint et course classique
def plot_sprint_vs_normal(df, post_id):
    sprint_distance = df['Distancesprint'].iloc[-1] if not df['Distancesprint'].empty and \
                                                            df['Distancesprint'].iloc[-1] != '-' else 0
    normal_distance = df['Distance'].iloc[-1] if not df['Distance'].empty and df['Distance'].iloc[-1] != '-' else 0

    if sprint_distance + normal_distance == 0:
        print(f"Aucune donnée de distance pour ID {post_id}")
        return None

    # Création du camembert
    plt.figure(figsize=(5, 5))
    labels = ['Sprint', 'Normal']
    values = [sprint_distance, normal_distance*1000]
    colors = ['red', 'green']
    plt.pie(values, labels=labels, autopct='%1.1f%%', colors=colors)
    plt.title(f"Sprint vs Normal (ID {post_id})")

    pie_path = f"sprint_normal_{post_id}.png"
    plt.savefig(pie_path)
    plt.close()

    return pie_path

#Génère le pdf
def generate_pdf_report(post_id):
    df = read_data_from_excel(post_id)
    if df is None:
        return

    speed_plot = plot_speed(df, post_id)
    sprint_pie = plot_sprint_vs_normal(df, post_id)

    # Résumés
    vmax = df['Vitessemax'].astype(float).max()
    distance = df['Distance'].iloc[-1] if not df['Distance'].empty else 0
    sprint_distance = df['Distancesprint'].iloc[-1] if not df['Distancesprint'].empty else 0
    nb_alertes = df['Alerte'].fillna('-').apply(lambda x: x != '-').sum()

    pdf = FPDF()
    pdf.add_page()

    pdf.set_font("Arial", size=16)
    pdf.cell(200, 10, f"Rapport Joueur ID {post_id}", ln=True, align='C')


    pdf.set_font("Arial", size=12)
    pdf.ln(10)
    pdf.cell(0, 10, f"Vitesse Max : {vmax} km/h", ln=True)
    pdf.cell(0, 10, f"Distance Totale : {distance} km", ln=True)
    pdf.cell(0, 10, f"Distance en Sprint : {sprint_distance} m", ln=True)
    pdf.cell(0, 10, f"Nombre d'alertes : {nb_alertes}", ln=True)

    pdf.ln(10)
    pdf.cell(0, 10, "Courbe de vitesse :", ln=True)
    pdf.image(speed_plot, w=180)

    pdf.ln(10)
    pdf.cell(0, 10, "Proportion Sprint vs Normal :", ln=True)
    pdf.image(sprint_pie, w=120)

    pdf_file = f"rapport_ID_{post_id}.pdf"
    pdf.output(pdf_file)
    print(f"PDF généré : {pdf_file}")

# Génère pour toutes les joueuses
def generate_all_reports():
    wb = load_workbook(EXCEL_FILE)
    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        if sheet_name.startswith("ID_"):
            post_id = sheet_name.replace("ID_", "")
            print(f"➡️ Génération du PDF pour l'ID {post_id}")
            generate_pdf_report(post_id)

    print("✅ Tous les rapports PDF ont été générés !")


#SERVER HTTP
@app.route("/")
def hello_world():
    return "<p> Hello World!</p>"

# Tu peux garder le message ici si utile pour autre chose, mais il n'est pas écrit dans Excel
@app.route('/postMessage/<int:postId>/<string:messageArduino>', methods=["POST"])
def post_message(postId, messageArduino):
    global message
    message = messageArduino
    return f'Message reçu : {message}'

@app.route('/post/alerte/<int:post_id>', methods=["POST"])
def post_alerte(post_id):
    liste_alerte.append(post_id)
    write_event(post_id, "Alerte", "Alerte")
    return f'Alerte {post_id}'

@app.route('/post/vitesse/<int:post_id>/<float:vitesse>', methods=["POST"])
def post_vitesse(post_id, vitesse):
    dic_vitesse[post_id] = vitesse
    write_event(post_id, "Vitesse", vitesse)
    return f'Vitesse enregistrée {post_id} -> {vitesse}'

@app.route('/post/distance/<int:post_id>/<float:distance>', methods=["POST"])
def post_distance(post_id, distance):
    dic_distance[post_id] = distance/1000 #pour mettre en km
    write_event(post_id, "Distance", distance/1000)
    return f'Distance enregistrée {post_id} -> {distance}'

@app.route('/post/vitesseMax/<int:post_id>/<float:vitesse>', methods=["POST"])
def post_vitesse_max(post_id, vitesse):
    dic_vitesse_max[post_id] = vitesse
    write_event(post_id, "VitesseMax", vitesse)
    return f'VitesseMax enregistrée {post_id} -> {vitesse}'

@app.route('/post/distanceSprint/<int:post_id>/<float:distance>', methods=["POST"])
def post_distance_sprint(post_id, distance):
    dic_distance_sprint[post_id] = int(distance)
    write_event(post_id, "DistanceSprint", int(distance))
    return f'DistanceSprint enregistrée {post_id} -> {distance}'

@app.route('/get/message')
def get_message():
    return jsonify(message)

@app.route('/get/alerte')
def get_alerte():
    global liste_alerte
    alerte_a_envoyer = liste_alerte.copy()
    liste_alerte.clear()
    return jsonify(alerte_a_envoyer)

@app.route('/get/vitesse')
def get_vitesse():
    return jsonify(dic_vitesse)

@app.route('/get/distance')
def get_distance():
    return jsonify(dic_distance)

@app.route('/get/vitesseMax')
def get_vitesse_max():
    return jsonify(dic_vitesse_max)

@app.route('/get/distanceSprint')
def get_distance_sprint():
    return jsonify(dic_distance_sprint)

@app.route('/export/pdf/all', methods=["GET"])
def export_all_pdfs():
    generate_all_reports()
    return "Tous les rapports PDF ont été générés !"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
